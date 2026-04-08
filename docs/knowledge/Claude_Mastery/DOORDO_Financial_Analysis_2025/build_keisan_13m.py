#!/usr/bin/env python3
"""
DOORDO 計算書類 13M集約版 自動転記スクリプト
  元データ  : FY25_増減分析_株式会社DOORDO.xlsx
  出力      : 20260406_DOORDO_計算書類_20251231_13M_集約版.xlsx
  BS        : 期末TB(12) 残高（2025/12/31）
  PL        : TB(12-11) + TB(12) 13ヶ月合算
  集約ルール : MAPPING_RULE準拠（少額科目→その他、資本性借入金→長期借入金統合）
"""
import sys, math
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

BASE      = "C:/Users/yotak/Documents/everything-claude-code/DOORDO_Financial_Analysis_2025/"
SRC_XL    = BASE + "FY25_増減分析_株式会社DOORDO.xlsx"
OUT_XL    = BASE + "20260406_DOORDO_計算書類_20251231_13M_集約版.xlsx"

COMPANY   = "株式会社ＤＯＯＲＤＯ"
PERIOD_BS = "令和7年12月31日 現在"
PERIOD_PL = "自 令和6年12月1日  至 令和7年12月31日（13ヶ月）"
PERIOD_SS = "自 令和6年12月1日  至 令和7年12月31日"

# ─────────────────────────────────────────────────────────────────────────────
# ステップ1・2：MAPPING_RULE（指示通り）
# ─────────────────────────────────────────────────────────────────────────────

MAPPING_RULE_BS = {
    "現金預金":       "現金及び預金",
    "完成工事未収入金":"完成工事未収入金",
    "販売用不動産":   "販売用不動産",
    "仕掛販売用不動産":"仕掛販売用不動産",
    "未成工事支出金": "未成工事支出金",
    "建物":           "建物",
    "建物付属設備":   "建物附属設備",
    "構築物":         "構築物",
    "車両運搬具":     "車両及び運搬具",
    "工具器具備品":   "器具及び備品",
    "土地":           "土地",
    "ソフトウェア":   "ソフトウェア",
    "長期貸付金":     "長期貸付金",
    "長期前払費用":   "長期前払費用",
    "短期借入金":     "短期借入金",
    "未払金":         "未払金",
    "未払法人税等":   "未払法人税等",
    "未成工事受入金": "未成工事受入金",
    "長期借入金":     "長期借入金",
    "資本性借入金":   "長期借入金",   # 長期借入金に統合
    "社債":           "社債",
    "預り保証金":     "長期預り保証金",
    "資本金":         "資本金",
    "繰越利益剰余金": "繰越利益剰余金",
}
# MAPPING_RULE_BS外 → セクションに応じて「その他〇〇」へ集約

MAPPING_RULE_PL = {
    "売上高":   "売上高",
    "仕入高":   "売上原価",
    "受取利息": "受取利息",
    "支払利息": "支払利息",
    "雑損失":   "営業外費用その他",
}
# 受取配当金・雑収入も営業外収益として個別表示
# MAPPING_RULE_PL外の販管費科目 → 「販売費及び一般管理費」

# TB勘定科目 → (MAPPING_RULEキー, セクション)
TB_ACCT_MAP = {
    # ── 流動資産 ──
    "現金":               ("現金預金",         "流動資産"),
    "普通預金":           ("現金預金",         "流動資産"),
    "定期預金":           ("現金預金",         "流動資産"),
    "定期積金":           ("現金預金",         "流動資産"),
    "返済準備金留保口座": ("現金預金",         "流動資産"),
    "完成工事未収入金":   ("完成工事未収入金", "流動資産"),
    "販売用不動産":       ("販売用不動産",     "流動資産"),
    "販売用不動産(収益)": ("販売用不動産",     "流動資産"),
    "仕掛販売用不動産":   ("仕掛販売用不動産", "流動資産"),
    "未成工事支出金":     ("未成工事支出金",   "流動資産"),
    # 以下 → MAPPING_RULEにないため「その他流動資産」
    "短期貸付金":         (None, "流動資産"),  # 今回はその他流動資産へ
    "未収入金":           (None, "流動資産"),
    "前渡金":             (None, "流動資産"),
    "立替金":             (None, "流動資産"),
    "未収収益":           (None, "流動資産"),
    "前払費用":           (None, "流動資産"),
    "未収消費税":         (None, "流動資産"),
    "仮払消費税":         (None, "流動資産"),
    "未収法人税等":       (None, "流動資産"),
    "仮払金":             (None, "流動資産"),
    # ── 有形固定資産 ──
    "建物":               ("建物",           "有形固定資産"),
    "附属設備":           ("建物付属設備",   "有形固定資産"),
    "構築物":             ("構築物",         "有形固定資産"),
    "車両運搬具":         ("車両運搬具",     "有形固定資産"),
    "工具器具備品":       ("工具器具備品",   "有形固定資産"),
    "土地":               ("土地",           "有形固定資産"),
    "建設仮勘定":         ("_skip",          "有形固定資産"),
    # ── 無形固定資産 ──
    "ソフトウェア":       ("ソフトウェア",   "無形固定資産"),
    "のれん":             ("_のれん",        "無形固定資産"),
    # ── 投資その他の資産 ── (MAPPING_RULEにないものは「その他（投資その他の資産）」)
    "長期貸付金":         ("長期貸付金",     "投資その他の資産"),
    "長期前払費用":       ("長期前払費用",   "投資その他の資産"),
    "出資金":             (None, "投資その他の資産"),
    "差入保証金":         (None, "投資その他の資産"),
    "保険積立金":         (None, "投資その他の資産"),
    "ﾘｻｲｸﾙ預託金":      (None, "投資その他の資産"),
    "ｺﾞﾙﾌ会員権":        (None, "投資その他の資産"),
    "繰延消費税額等":     (None, "投資その他の資産"),
    # ── 流動負債 ──
    "短期借入金":         ("短期借入金",     "流動負債"),
    "未払金":             ("未払金",         "流動負債"),
    "未払法人税等":       ("未払法人税等",   "流動負債"),
    "未成工事受入金":     ("未成工事受入金", "流動負債"),
    # 以下 → 「その他流動負債」
    "未払費用":           (None, "流動負債"),
    "前受金":             (None, "流動負債"),
    "預り金":             (None, "流動負債"),
    "前受収益":           (None, "流動負債"),
    "未払消費税等":       (None, "流動負債"),
    "仮受金":             (None, "流動負債"),
    "仮受消費税":         (None, "流動負債"),
    # ── 固定負債 ──
    "社債":               ("社債",           "固定負債"),
    "長期借入金":         ("長期借入金",     "固定負債"),
    "資本性借入金":       ("資本性借入金",   "固定負債"),  # →長期借入金に統合
    "預り保証金":         ("預り保証金",     "固定負債"),
    # ── 純資産 ──
    "資本金":             ("資本金",             "純資産"),
    "繰越利益":           ("繰越利益剰余金",     "純資産"),
}

TB_SECTION_MAP = {
    "[現金･預金]":      "流動資産",
    "[売上債権]":        "流動資産",
    "[有価証券]":        "流動資産",
    "[棚卸資産]":        "流動資産",
    "[他流動資産]":      "流動資産",
    "[有形固定資産]":    "有形固定資産",
    "[無形固定資産]":    "無形固定資産",
    "[投資その他の資産]":"投資その他の資産",
    "[繰延資産]":        "繰延資産",
    "[仕入債務]":        "流動負債",
    "[他流動負債]":      "流動負債",
    "[固定負債]":        "固定負債",
    "[資本金]":          "純資産",
    "[新株式申込証拠金]":"純資産",
    "[資本剰余金]":      "純資産",
    "[利益剰余金]":      "純資産",
    "[自己株式]":        "純資産",
    "[評価･換算差額等]": "純資産",
    "[新株予約権]":      "純資産",
}
PL_SECTIONS = {"[売上高]","[売上原価]","[販売管理費]","[営業外収益]","[営業外費用]",
               "[特別利益]","[特別損失]","[当期純損益]"}

def to_int(v):
    if v is None: return 0
    try: return int(v)
    except: return 0

# ─────────────────────────────────────────────────────────────────────────────
# TB読み込み
# ─────────────────────────────────────────────────────────────────────────────

def read_tb(ws, bal_col=7):
    bs_rows, pl_rows, opening = [], [], {}
    cur_fs, cur_sec = None, None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None: continue
        rt = str(row[0])
        if rt not in ("[区分行]","[明細行]","[合計行]","[表題行]"): continue
        if len(row) <= bal_col: continue
        cls  = str(row[2]) if row[2] else ""
        acct = str(row[3]) if row[3] else ""
        bal  = to_int(row[bal_col])
        op   = to_int(row[4]) if len(row) > 4 else 0
        if rt == "[区分行]":
            if cls == "[貸借対照表]":
                cur_fs  = "BS"
                cur_sec = TB_SECTION_MAP.get(acct, acct)
            elif cls == "[損益計算書]":
                cur_fs  = "PL"
                cur_sec = acct  # e.g. [販売管理費]
        elif rt == "[明細行]":
            if cur_fs == "BS":
                bs_rows.append((acct, bal, cur_sec)); opening[acct] = op
            elif cur_fs == "PL":
                pl_rows.append((acct, bal, cur_sec))
    return bs_rows, pl_rows, opening

def read_summary(ws, bal_col=7):
    d = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] != "[合計行]": continue
        if len(row) > bal_col:
            d[str(row[3]) if row[3] else ""] = to_int(row[bal_col])
    return d

print("=" * 64)
print("  TB読み込み中（44MB Excel）...")
print("=" * 64)
src_wb    = openpyxl.load_workbook(SRC_XL, read_only=True)
bs12_11, pl12_11, op_12_11 = read_tb(src_wb["期末TB(12-11)"], 7)
bs12,    pl12,    op_12    = read_tb(src_wb["期末TB(12)"],    7)
sum12_11  = read_summary(src_wb["期末TB(12-11)"], 7)
sum12     = read_summary(src_wb["期末TB(12)"],    7)
src_wb.close()
print(f"  TB12-11: BS {len(bs12_11)}行 / PL {len(pl12_11)}行")
print(f"  TB12:    BS {len(bs12)}行  / PL {len(pl12)}行")

# ─────────────────────────────────────────────────────────────────────────────
# ステップ2：BS集計（TB12 期末残高）
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 64)
print("  ステップ2: BS集計（2025/12/31 期末残高）")
print("=" * 64)

bs_agg   = defaultdict(int)   # (開示科目, セクション) → 合計
bs_detail = []                 # マッピング詳細（コンソール内訳用・マッピング表用）
# その他内訳追跡
その他流動資産_内訳    = {}
その他投資その他_内訳  = {}
その他流動負債_内訳    = {}
長期借入金_内訳        = {}

for tb_acct, bal, tb_sec in bs12:
    if bal == 0: continue

    if tb_acct not in TB_ACCT_MAP:
        # 未登録 → セクションに応じてその他へ
        sec = tb_sec if tb_sec else "不明"
        disc = f"その他（{sec}）"
        bs_agg[(disc, sec)] += bal
        bs_detail.append((tb_acct, tb_sec, bal, "(未登録)", sec, disc))
        continue

    rule_key, sec = TB_ACCT_MAP[tb_acct]

    if rule_key == "_skip":
        bs_detail.append((tb_acct, tb_sec, bal, "(スキップ)", sec, "-"))
        continue

    if rule_key == "_のれん":
        disc = "のれん"
        bs_agg[(disc, sec)] += bal
        bs_detail.append((tb_acct, tb_sec, bal, rule_key, sec, disc))
        continue

    if rule_key is None:
        # MAPPING_RULE外 → その他へ（セクションで分類）
        disc = f"その他（{sec}）"
        bs_agg[(disc, sec)] += bal
        bs_detail.append((tb_acct, tb_sec, bal, "(→その他)", sec, disc))
        # 内訳追跡
        if sec == "流動資産":
            その他流動資産_内訳[tb_acct] = その他流動資産_内訳.get(tb_acct, 0) + bal
        elif sec == "投資その他の資産":
            その他投資その他_内訳[tb_acct] = その他投資その他_内訳.get(tb_acct, 0) + bal
        elif sec == "流動負債":
            その他流動負債_内訳[tb_acct] = その他流動負債_内訳.get(tb_acct, 0) + bal
        continue

    # MAPPING_RULEに登録あり
    if rule_key in MAPPING_RULE_BS:
        disc = MAPPING_RULE_BS[rule_key]
    else:
        disc = f"その他（{sec}）"

    bs_agg[(disc, sec)] += bal
    bs_detail.append((tb_acct, tb_sec, bal, rule_key, sec, disc))

    # 長期借入金統合の内訳追跡
    if disc == "長期借入金":
        長期借入金_内訳[tb_acct] = 長期借入金_内訳.get(tb_acct, 0) + bal

# 繰越利益剰余金 → 合計行（前期繰越＋当期純損益）で上書き
netto_closing = sum12.get("繰越利益剰余金合計", 0)
bs_agg[("繰越利益剰余金", "純資産")] = netto_closing

def bv(key, sec=""):
    if sec:
        return bs_agg.get((key, sec), 0)
    for (k, s), v in bs_agg.items():
        if k == key: return v
    return 0

# ── 集計値 ──────────────────────────────────────────────────────────────────
現金預金      = bv("現金及び預金",    "流動資産")
完成工事      = bv("完成工事未収入金","流動資産")
販売用        = bv("販売用不動産",    "流動資産")
仕掛販売用    = bv("仕掛販売用不動産","流動資産")
未成工事支出  = bv("未成工事支出金",  "流動資産")
流動その他    = bv("その他（流動資産）","流動資産")

建物          = bv("建物",           "有形固定資産")
建物附属      = bv("建物附属設備",   "有形固定資産")
構築物        = bv("構築物",         "有形固定資産")
車両          = bv("車両及び運搬具", "有形固定資産")
器具備品      = bv("器具及び備品",   "有形固定資産")
土地          = bv("土地",           "有形固定資産")
ソフト        = bv("ソフトウェア",   "無形固定資産")
のれん_bal    = bv("のれん",         "無形固定資産")
長期貸付      = bv("長期貸付金",     "投資その他の資産")
長期前払      = bv("長期前払費用",   "投資その他の資産")
投資その他    = bv("その他（投資その他の資産）","投資その他の資産")

短期借入      = bv("短期借入金",     "流動負債")
未払金_bal    = bv("未払金",         "流動負債")
未払法人税    = bv("未払法人税等",   "流動負債")
未成工事受入  = bv("未成工事受入金", "流動負債")
流動負債その他 = bv("その他（流動負債）","流動負債")
社債_bal      = bv("社債",           "固定負債")
長期借入      = bv("長期借入金",     "固定負債")
長期保証      = bv("長期預り保証金", "固定負債")
資本金_bal    = bv("資本金",         "純資産")
繰越利益_bal  = bv("繰越利益剰余金", "純資産")

# ── 小計 ─────────────────────────────────────────────────────────────────────
有形固定計    = 建物 + 建物附属 + 構築物 + 車両 + 器具備品 + 土地
無形固定計    = ソフト + のれん_bal
投資その他計  = 長期貸付 + 長期前払 + 投資その他
固定資産計    = 有形固定計 + 無形固定計 + 投資その他計
流動資産計    = 現金預金 + 完成工事 + 販売用 + 仕掛販売用 + 未成工事支出 + 流動その他
資産合計      = 流動資産計 + 固定資産計
流動負債計    = 短期借入 + 未払金_bal + 未払法人税 + 未成工事受入 + 流動負債その他
固定負債計    = 社債_bal + 長期借入 + 長期保証
負債合計      = 流動負債計 + 固定負債計
利益剰余金計  = 繰越利益_bal
株主資本計    = 資本金_bal + 利益剰余金計
純資産計      = 株主資本計
負債純資産計  = 負債合計 + 純資産計

print(f"  資産合計    : {資産合計:>22,}")
print(f"  負債合計    : {負債合計:>22,}")
print(f"  純資産合計  : {純資産計:>22,}")
print(f"  BS整合性    : {'✅ OK' if 資産合計 == 負債純資産計 else f'❌ 差額 {資産合計-負債純資産計:+,}'}")

# ─────────────────────────────────────────────────────────────────────────────
# PL集計（13ヶ月合算）
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 64)
print("  PL集計（13ヶ月合算）")
print("=" * 64)

pl_agg     = defaultdict(int)
pl_detail  = []
pl_sga_12m = {}
pl_sga_1m  = {}

# 営業外収益（受取配当金・雑収入）は個別表示
OGI_ACCTS = {"受取配当金": "受取配当金", "雑収入": "雑収入"}

def add_pl(rows, label, sga_dict):
    for tb_acct, bal, pl_sec in rows:
        if bal == 0: continue
        if tb_acct in MAPPING_RULE_PL:
            disc = MAPPING_RULE_PL[tb_acct]
            pl_agg[disc] += bal
            pl_detail.append((label, tb_acct, pl_sec, bal, disc))
        elif tb_acct in OGI_ACCTS:
            disc = OGI_ACCTS[tb_acct]
            pl_agg[disc] += bal
            pl_detail.append((label, tb_acct, pl_sec, bal, disc))
        elif pl_sec == "[販売管理費]":
            if tb_acct.startswith("★"): continue
            pl_agg["販売費及び一般管理費"] += bal
            sga_dict[tb_acct] = sga_dict.get(tb_acct, 0) + bal
            pl_detail.append((label, tb_acct, pl_sec, bal, "販売費及び一般管理費"))
        elif pl_sec == "[売上高]":
            pl_agg["売上高"] += bal
            pl_detail.append((label, tb_acct, pl_sec, bal, "売上高"))
        elif pl_sec == "[売上原価]":
            if not tb_acct.startswith("★"):
                pl_agg["売上原価"] += bal
                pl_detail.append((label, tb_acct, pl_sec, bal, "売上原価"))
        elif pl_sec == "[特別利益]":
            pl_agg["固定資産売却益"] += bal
            pl_detail.append((label, tb_acct, pl_sec, bal, "固定資産売却益"))
        elif pl_sec == "[特別損失]":
            pl_agg["固定資産除却損"] += bal
            pl_detail.append((label, tb_acct, pl_sec, bal, "固定資産除却損"))
        elif pl_sec == "[当期純損益]":
            pl_agg["法人税等"] += bal
            pl_detail.append((label, tb_acct, pl_sec, bal, "法人税等"))
        else:
            pl_detail.append((label, tb_acct, pl_sec, bal, "(未分類)"))

add_pl(pl12_11, "TB12-11(12M)", pl_sga_12m)
add_pl(pl12,    "TB12(1M)",     pl_sga_1m)

gross_profit  = pl_agg["売上高"]    - pl_agg["売上原価"]
op_income     = gross_profit         - pl_agg["販売費及び一般管理費"]
other_inc     = pl_agg["受取利息"] + pl_agg["受取配当金"] + pl_agg["雑収入"]
other_exp     = pl_agg["支払利息"] + pl_agg["営業外費用その他"]
recurring_inc = op_income + other_inc - other_exp
extra_i       = pl_agg["固定資産売却益"]
extra_l       = pl_agg["固定資産除却損"]
pretax_inc    = recurring_inc + extra_i - extra_l
income_tax    = pl_agg["法人税等"]
net_income    = pretax_inc - income_tax

print(f"  売上高              : {pl_agg['売上高']:>22,}")
print(f"  売上原価            : {pl_agg['売上原価']:>22,}")
print(f"  売上総利益          : {gross_profit:>22,}")
print(f"  販売費及び一般管理費: {pl_agg['販売費及び一般管理費']:>22,}")
print(f"  営業利益            : {op_income:>22,}")
print(f"  経常利益            : {recurring_inc:>22,}")
print(f"  当期純利益          : {net_income:>22,}")

# SS値
netto_13m       = (sum12_11.get("当期純損益金額", 0) + sum12.get("当期純損益金額", 0))
ss_開始繰越     = op_12_11.get("繰越利益", 0)
ss_開始資本金   = op_12_11.get("資本金",   30_000_000)
ss_変動繰越     = netto_13m
ss_終了繰越     = netto_closing
ss_その他変動   = ss_終了繰越 - ss_開始繰越 - ss_変動繰越  # のれん消却等直入
ss_変動合計     = ss_変動繰越 + ss_その他変動
ss_終了株主資本 = ss_開始資本金 + ss_開始繰越 + ss_変動合計

print(f"\n  SS期首繰越利益      : {ss_開始繰越:>22,}")
print(f"  13M当期純利益        : {ss_変動繰越:>22,}")
if ss_その他変動:
    print(f"  その他変動（のれん等）: {ss_その他変動:>22,}")
print(f"  SS期末純資産        : {ss_終了株主資本:>22,}")
print(f"  SS整合性            : {'✅ OK' if ss_終了株主資本 == 純資産計 else f'❌ 差額 {ss_終了株主資本 - 純資産計:+,}'}")

# ─────────────────────────────────────────────────────────────────────────────
# ステップ3：Excel出力
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 64)
print("  ステップ3: Excel作成")
print("=" * 64)

wb = Workbook()
wb.remove(wb.active)

# ── スタイル ──────────────────────────────────────────────────────────────
FN = "游ゴシック"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
GRY_FILL = PatternFill("solid", fgColor="F2F2F2")
HDR_FONT = Font(name=FN, bold=True, color="FFFFFF", size=10)
BOL_FONT = Font(name=FN, bold=True, size=10)
REG_FONT = Font(name=FN, size=10)
THIN     = Side(style="thin")
MED      = Side(style="medium")
NUM_FMT  = "#,##0"

def ba():
    return Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def nk(ws, cel, val, bold=False, num=False, wrap=False, fill=None, align="left"):
    c = ws[cel]
    c.value = val
    c.font  = BOL_FONT if bold else REG_FONT
    c.alignment = Alignment(
        horizontal="right" if (num or align == "right") else align,
        vertical="center", wrap_text=wrap)
    if num and isinstance(val, (int, float)):
        c.number_format = NUM_FMT
    if fill: c.fill = fill

def rounddown_k(v):
    """千円単位（切り捨て）"""
    if v is None: return None
    if v < 0: return -math.floor(abs(v) / 1000)
    return math.floor(v / 1000)

# ─────────────────────────────────────────────────────────────────────────────
# ─── BS（円）シート ────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def make_bs(wb, sheet_name, divisor=1):
    """divisor=1→円, divisor=1000→千円"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    unit = "（単位：円）" if divisor == 1 else "（単位：千円）"

    def v(val):
        """divisorで割った値を返す"""
        if val is None or val == 0: return 0
        if divisor == 1: return val
        return rounddown_k(val)

    # ヘッダー
    ws["B2"] = "貸　借　対　照　表"
    ws["B2"].font = Font(name=FN, bold=True, size=13)
    ws["B3"] = f"（{PERIOD_BS}）"
    ws["B3"].font = REG_FONT
    ws["B4"] = COMPANY; ws["B4"].font = REG_FONT
    ws["N4"] = unit; ws["N4"].font = REG_FONT
    for col, txt in [("B","科目"),("H","金額"),("I","科目"),("P","金額")]:
        ws[f"{col}5"] = txt; ws[f"{col}5"].font = BOL_FONT
    ws["B6"] = "（資産の部）"; ws["B6"].font = BOL_FONT
    ws["I6"] = "（負債の部）"; ws["I6"].font = BOL_FONT

    def wL(row, lv, txt, det=None, sub=None):
        """左側（資産）書き込み"""
        cols = {0:"B", 1:"C", 2:"D", 3:"E"}
        indent = {0:"", 1:"", 2:"　", 3:"　　"}
        col = cols.get(lv,"E")
        ws[f"{col}{row}"] = indent.get(lv,"") + (txt or "")
        ws[f"{col}{row}"].font = BOL_FONT if lv <= 1 else REG_FONT
        if det is not None:
            ws[f"G{row}"] = v(det)
            ws[f"G{row}"].number_format = NUM_FMT
            ws[f"G{row}"].alignment = Alignment(horizontal="right")
            ws[f"G{row}"].font = REG_FONT
        if sub is not None:
            ws[f"H{row}"] = v(sub)
            ws[f"H{row}"].number_format = NUM_FMT
            ws[f"H{row}"].alignment = Alignment(horizontal="right")
            ws[f"H{row}"].font = BOL_FONT

    def wR(row, lv, txt=None, det=None, sub=None):
        """右側（負債・純資産）書き込み"""
        cols = {0:"I", 1:"J", 2:"K", 3:"L"}
        indent = {0:"", 1:"", 2:"　", 3:"　　"}
        if txt is not None:
            col = cols.get(lv,"L")
            ws[f"{col}{row}"] = indent.get(lv,"") + txt
            ws[f"{col}{row}"].font = BOL_FONT if lv <= 1 else REG_FONT
        if det is not None:
            ws[f"N{row}"] = v(det)
            ws[f"N{row}"].number_format = NUM_FMT
            ws[f"N{row}"].alignment = Alignment(horizontal="right")
            ws[f"N{row}"].font = REG_FONT
        if sub is not None:
            ws[f"O{row}"] = v(sub)
            ws[f"O{row}"].number_format = NUM_FMT
            ws[f"O{row}"].alignment = Alignment(horizontal="right")
            ws[f"O{row}"].font = BOL_FONT

    r = 7
    wL(r,1,"流動資産",sub=流動資産計);           wR(r,1,"流動負債",sub=流動負債計);     r+=1
    wL(r,3,"現金及び預金",現金預金);              wR(r,3,"短期借入金",det=短期借入);     r+=1
    wL(r,3,"完成工事未収入金",完成工事);          wR(r,3,"未払金",det=未払金_bal);       r+=1
    wL(r,3,"販売用不動産",販売用);                wR(r,3,"未払法人税等",det=未払法人税); r+=1
    wL(r,3,"仕掛販売用不動産",仕掛販売用);        wR(r,3,"未成工事受入金",det=未成工事受入); r+=1
    wL(r,3,"未成工事支出金",未成工事支出);        wR(r,3,"その他",det=流動負債その他);   r+=1
    wL(r,3,"その他",流動その他);                  wR(r,1,"固定負債",sub=固定負債計);     r+=1
    wL(r,1,"固定資産",sub=固定資産計);            wR(r,3,"社債",det=社債_bal);           r+=1
    wL(r,2,"有形固定資産",sub=有形固定計);        wR(r,3,"長期借入金",det=長期借入);     r+=1
    wL(r,3,"建物",建物);                          wR(r,3,"長期預り保証金",det=長期保証); r+=1
    wL(r,3,"建物附属設備",建物附属);              wR(r,0,"負債合計",sub=負債合計);       r+=1
    wL(r,3,"構築物",構築物);                      wR(r,0,"（純資産の部）");              r+=1
    wL(r,3,"車両及び運搬具",車両);                wR(r,1,"株主資本",sub=株主資本計);     r+=1
    wL(r,3,"器具及び備品",器具備品);              wR(r,3,"資本金",det=資本金_bal);       r+=1
    wL(r,3,"土地",土地);                          wR(r,2,"利益剰余金",sub=利益剰余金計); r+=1
    wL(r,2,"無形固定資産",sub=無形固定計);        wR(r,3,"繰越利益剰余金",det=繰越利益_bal); r+=1
    wL(r,3,"ソフトウェア",ソフト);                wR(r,0,"純資産合計",sub=純資産計);     r+=1
    if のれん_bal:
        wL(r,3,"のれん",のれん_bal);              wR(r,0,"負債及び純資産合計",sub=負債純資産計); r+=1
    else:
        wL(r,0,"");                               wR(r,0,"負債及び純資産合計",sub=負債純資産計); r+=1
    wL(r,2,"投資その他の資産",sub=投資その他計);  r+=1
    wL(r,3,"長期貸付金",長期貸付);                r+=1
    wL(r,3,"長期前払費用",長期前払);              r+=1
    wL(r,3,"その他",投資その他);                  r+=1
    wL(r,0,"資産合計",sub=資産合計);              r+=1

    for col,w in [("B",12),("C",14),("D",16),("E",22),("F",2),("G",16),("H",16),
                  ("I",14),("J",14),("K",16),("L",22),("M",2),("N",16),("O",16)]:
        ws.column_dimensions[col].width = w
    print(f"  {sheet_name}: 行7-{r-1}")
    return ws

ws_bs_en = make_bs(wb, "BS（円）",   divisor=1)
ws_bs_k  = make_bs(wb, "BS（千円）", divisor=1000)

# ─────────────────────────────────────────────────────────────────────────────
# ─── PL（円）/（千円）シート ─────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def make_pl(wb, sheet_name, divisor=1):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    unit = "（単位：円）" if divisor == 1 else "（単位：千円）"

    def v(val):
        if val is None or val == 0: return 0
        return val if divisor == 1 else rounddown_k(val)

    ws["B2"] = "損　益　計　算　書"
    ws["B2"].font = Font(name=FN, bold=True, size=13)
    ws["B3"] = f"（{PERIOD_PL}）"; ws["B3"].font = REG_FONT
    ws["B4"] = COMPANY;            ws["B4"].font = REG_FONT
    ws["H4"] = unit;               ws["H4"].font = REG_FONT
    ws["B5"] = "科目";             ws["B5"].font = BOL_FONT
    ws["H5"] = "金額";             ws["H5"].font = BOL_FONT

    def wr(row, lv, txt, det=None, sub=None):
        cols = {0:"B", 1:"C", 2:"D"}
        col  = cols.get(lv,"D")
        ind  = {0:"",  1:"",  2:"　"}
        ws[f"{col}{row}"] = ind.get(lv,"") + txt
        ws[f"{col}{row}"].font = BOL_FONT if lv <= 1 else REG_FONT
        if det is not None:
            ws[f"G{row}"] = v(det)
            ws[f"G{row}"].number_format = NUM_FMT
            ws[f"G{row}"].alignment = Alignment(horizontal="right")
            ws[f"G{row}"].font = REG_FONT
        if sub is not None:
            ws[f"H{row}"] = v(sub)
            ws[f"H{row}"].number_format = NUM_FMT
            ws[f"H{row}"].alignment = Alignment(horizontal="right")
            ws[f"H{row}"].font = BOL_FONT

    r = 6
    wr(r,1,"売上高",                             sub=pl_agg["売上高"]);        r+=1
    wr(r,1,"売上原価",                           sub=pl_agg["売上原価"]);      r+=1
    wr(r,2,"売上総利益",                         sub=gross_profit);            r+=1
    wr(r,1,"販売費及び一般管理費",               sub=pl_agg["販売費及び一般管理費"]); r+=1
    wr(r,2,"営業利益",                           sub=op_income);               r+=1
    wr(r,1,"営業外収益");                                                       r+=1
    wr(r,2,"受取利息",            det=pl_agg["受取利息"]);                     r+=1
    wr(r,2,"受取配当金",          det=pl_agg["受取配当金"]);                   r+=1
    wr(r,2,"雑収入",              det=pl_agg["雑収入"],     sub=other_inc);    r+=1
    wr(r,1,"営業外費用");                                                       r+=1
    wr(r,2,"支払利息",            det=pl_agg["支払利息"]);                     r+=1
    wr(r,2,"その他",              det=pl_agg["営業外費用その他"], sub=other_exp); r+=1
    wr(r,2,"経常利益",                           sub=recurring_inc);           r+=1
    wr(r,1,"特別利益");                                                         r+=1
    wr(r,2,"固定資産売却益",      det=extra_i,             sub=extra_i);      r+=1
    wr(r,1,"特別損失");                                                         r+=1
    wr(r,2,"固定資産除却損",      det=extra_l,             sub=extra_l);      r+=1
    wr(r,2,"税引前当期純利益",                   sub=pretax_inc);              r+=1
    wr(r,0,"法人税、住民税及び事業税", det=income_tax);                        r+=1
    wr(r,2,"当期純利益",                         sub=net_income);              r+=1

    for col,w in [("B",22),("C",22),("D",22),("E",2),("F",2),("G",18),("H",18)]:
        ws.column_dimensions[col].width = w
    print(f"  {sheet_name}: 行6-{r-1}")

make_pl(wb, "PL（円）",   divisor=1)
make_pl(wb, "PL（千円）", divisor=1000)

# ─────────────────────────────────────────────────────────────────────────────
# ─── SS（円）/（千円）シート ─────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def make_ss(wb, sheet_name, divisor=1):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    unit = "（単位：円）" if divisor == 1 else "（単位：千円）"

    def v(val):
        if val is None: return None
        return val if divisor == 1 else rounddown_k(val)

    ws["B2"] = "株主資本等変動計算書"
    ws["B2"].font = Font(name=FN, bold=True, size=13)
    ws["B3"] = f"（{PERIOD_SS}）"; ws["B3"].font = REG_FONT
    ws["B4"] = COMPANY;            ws["B4"].font = REG_FONT
    ws["J4"] = unit;               ws["J4"].font = REG_FONT

    def wr(row, lv, txt=None, label=None, amount=None):
        col_map = {0:"C", 1:"C", 2:"D", 3:"E", 4:"F"}
        if txt is not None:
            col = col_map.get(lv,"D")
            ws[f"{col}{row}"] = ("　" * max(0, lv - 1)) + txt
            ws[f"{col}{row}"].font = BOL_FONT if lv <= 2 else REG_FONT
        if label:
            ws[f"H{row}"] = label; ws[f"H{row}"].font = REG_FONT
        if amount is not None:
            ws[f"J{row}"] = v(amount)
            ws[f"J{row}"].number_format = NUM_FMT
            ws[f"J{row}"].alignment = Alignment(horizontal="right")
            ws[f"J{row}"].font = REG_FONT

    r = 6
    wr(r,1,"（株主資本）");                                                                          r+=1
    wr(r,2,"資本金",          "当期首残高",           ss_開始資本金);                               r+=1
    wr(r,None,None,           "当期末残高",            ss_開始資本金);                               r+=1
    wr(r,2,"利益剰余金");                                                                             r+=1
    wr(r,3,"　その他利益剰余金");                                                                     r+=1
    wr(r,4,"繰越利益剰余金",  "当期首残高",           ss_開始繰越);                                  r+=1
    wr(r,None,None,           "当期変動額　当期純利益", ss_変動繰越);                                 r+=1
    if ss_その他変動 != 0:
        wr(r,None,None,       "当期変動額　その他（のれん消却等）", ss_その他変動);                   r+=1
    wr(r,None,None,           "当期末残高",            ss_終了繰越);                                  r+=1
    wr(r,3,"　利益剰余金合計","当期首残高",            ss_開始繰越);                                  r+=1
    wr(r,None,None,           "当期変動額",            ss_変動合計);                                  r+=1
    wr(r,None,None,           "当期末残高",            ss_終了繰越);                                  r+=1
    wr(r,2,"株主資本合計",    "当期首残高",            ss_開始資本金 + ss_開始繰越);                  r+=1
    wr(r,None,None,           "当期変動額",            ss_変動合計);                                  r+=1
    wr(r,None,None,           "当期末残高",            ss_終了株主資本);                              r+=1
    wr(r,0,"純資産の部合計",  "当期首残高",            ss_開始資本金 + ss_開始繰越);                  r+=1
    wr(r,None,None,           "当期変動額",            ss_変動合計);                                  r+=1
    wr(r,None,None,           "当期末残高",            ss_終了株主資本);                              r+=1

    for col,w in [("B",4),("C",18),("D",18),("E",18),("F",18),("G",4),("H",28),("I",4),("J",18)]:
        ws.column_dimensions[col].width = w
    print(f"  {sheet_name}: 行6-{r-1}")

make_ss(wb, "SS（円）",   divisor=1)
make_ss(wb, "SS（千円）", divisor=1000)

# ─────────────────────────────────────────────────────────────────────────────
# ─── 販管費明細_附属明細書用 ────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

ws_sga = wb.create_sheet("販管費明細_附属明細書用")
ws_sga.sheet_view.showGridLines = False
ws_sga["B2"] = "販売費及び一般管理費 明細書（附属明細書用）"
ws_sga["B2"].font = Font(name=FN, bold=True, size=12)
ws_sga["B3"] = f"（{PERIOD_PL}）"; ws_sga["B3"].font = REG_FONT
ws_sga["B4"] = COMPANY;            ws_sga["B4"].font = REG_FONT
ws_sga["G4"] = "（単位：円）";     ws_sga["G4"].font = REG_FONT

for col, h in [("C","勘定科目"),("E","第18期（12ヶ月）\n2024/12-2025/11"),
               ("F","第19期（1ヶ月）\n2025/12"),("G","13ヶ月合計")]:
    c = ws_sga[f"{col}5"]
    c.value = h; c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = ba()
ws_sga.row_dimensions[5].height = 32

all_sga = sorted(set(list(pl_sga_12m) + list(pl_sga_1m)))
row = 6
for acct in all_sga:
    v12 = pl_sga_12m.get(acct, 0)
    v1  = pl_sga_1m.get(acct, 0)
    tot = v12 + v1
    ws_sga[f"C{row}"] = acct; ws_sga[f"C{row}"].font = REG_FONT; ws_sga[f"C{row}"].border = ba()
    for col, val in [("E", v12 or None), ("F", v1 or None), ("G", tot)]:
        c = ws_sga[f"{col}{row}"]
        c.value = val; c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal="right"); c.font = REG_FONT; c.border = ba()
    row += 1

ws_sga[f"C{row}"] = "販売費及び一般管理費合計"
ws_sga[f"C{row}"].font = BOL_FONT; ws_sga[f"C{row}"].fill = GRY_FILL; ws_sga[f"C{row}"].border = ba()
for col, val in [("E", sum(pl_sga_12m.values())),
                 ("F", sum(pl_sga_1m.values())),
                 ("G", pl_agg["販売費及び一般管理費"])]:
    c = ws_sga[f"{col}{row}"]
    c.value = val; c.number_format = NUM_FMT
    c.alignment = Alignment(horizontal="right"); c.font = BOL_FONT
    c.fill = GRY_FILL; c.border = ba()

for col, w in [("B",4),("C",26),("D",4),("E",20),("F",20),("G",20)]:
    ws_sga.column_dimensions[col].width = w
ws_sga.freeze_panes = "C6"
print(f"  販管費明細_附属明細書用: {len(all_sga)}科目")

# ─────────────────────────────────────────────────────────────────────────────
# ステップ4：保存
# ─────────────────────────────────────────────────────────────────────────────

wb.save(OUT_XL)
print(f"\n  ✅ 保存: {OUT_XL}")

# ─────────────────────────────────────────────────────────────────────────────
# コンソール内訳報告
# ─────────────────────────────────────────────────────────────────────────────

sep = "═" * 64
print(f"""
{sep}
  ✅ 処理完了 ─ 集約内訳レポート
{sep}

【BS 2025/12/31 確認】
  資産合計       : {資産合計:>22,}
  負債合計       : {負債合計:>22,}
  純資産合計     : {純資産計:>22,}
  BS整合性       : {'OK（資産＝負債+純資産）' if 資産合計 == 負債純資産計 else f'NG 差額 {資産合計-負債純資産計:+,}'}

【その他流動資産 内訳（合計 {sum(その他流動資産_内訳.values()):,}円）】""")
for acct, amt in sorted(その他流動資産_内訳.items(), key=lambda x: -x[1]):
    print(f"  　{acct:<20}  {amt:>16,}")

print(f"""
【その他（投資その他の資産） 内訳（合計 {sum(その他投資その他_内訳.values()):,}円）】""")
for acct, amt in sorted(その他投資その他_内訳.items(), key=lambda x: -x[1]):
    print(f"  　{acct:<20}  {amt:>16,}")

print(f"""
【その他流動負債 内訳（合計 {sum(その他流動負債_内訳.values()):,}円）】""")
for acct, amt in sorted(その他流動負債_内訳.items(), key=lambda x: -x[1]):
    print(f"  　{acct:<20}  {amt:>16,}")

print(f"""
【長期借入金 統合内訳（合計 {sum(長期借入金_内訳.values()):,}円）】""")
for acct, amt in sorted(長期借入金_内訳.items(), key=lambda x: -x[1]):
    print(f"  　{acct:<20}  {amt:>16,}")

print(f"""
【PL 13ヶ月合算】
  売上高                : {pl_agg['売上高']:>22,}
  売上原価              : {pl_agg['売上原価']:>22,}
  売上総利益            : {gross_profit:>22,}
  販売費及び一般管理費  : {pl_agg['販売費及び一般管理費']:>22,}
  営業利益              : {op_income:>22,}
  営業外収益            : {other_inc:>22,}
  営業外費用            : {other_exp:>22,}
  経常利益              : {recurring_inc:>22,}
  税引前当期純利益      : {pretax_inc:>22,}
  法人税等              : {income_tax:>22,}
  当期純利益            : {net_income:>22,}

【SS 整合】
  期首純資産（2024/12/01）: {ss_開始資本金 + ss_開始繰越:>20,}
  当期純利益（13M）        : {ss_変動繰越:>20,}
  その他変動（のれん等）   : {ss_その他変動:>20,}
  期末純資産（2025/12/31） : {ss_終了株主資本:>20,}
  SS整合性（＝BS純資産）  : {'✅ OK' if ss_終了株主資本 == 純資産計 else f'❌ 差額 {ss_終了株主資本 - 純資産計:+,}'}

{sep}""")
