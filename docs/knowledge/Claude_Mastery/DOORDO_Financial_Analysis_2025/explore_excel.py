import openpyxl

# Use read_only=True for faster loading of large file
wb = openpyxl.load_workbook(
    'FY25_増減分析_株式会社DOORDO.xlsx',
    data_only=True,
    read_only=True
)
sheet_names = wb.sheetnames

lines = []
lines.append('=== Sheet Names ===')
for i, s in enumerate(sheet_names):
    lines.append(f'{i}: {s}')

lines.append('')

# Scan each sheet for key financial terms
keywords = ['売上高', '当期純', '純利益', '資産合計', '純資産', '売上', '合計']

for sheet_name in sheet_names:
    ws = wb[sheet_name]
    found_rows = []
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if row_count > 500:
            break
        for cell in row:
            if isinstance(cell, str):
                for kw in keywords:
                    if kw in cell:
                        found_rows.append(f'  row{row_count}: {str(row[:10])}')
                        break
    if found_rows:
        lines.append(f'--- Sheet: {sheet_name} (keyword hits) ---')
        lines.extend(found_rows[:15])
        lines.append('')

wb.close()

with open('xl_out.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
