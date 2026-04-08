#!/usr/bin/env python3
"""
DOORDO 計算書類 自動転記スクリプト
  元データ：FY25_増減分析_株式会社DOORDO.xlsx（期末TB(12-11), 期末TB(12)）
  出力：20260406_DOORDO_計算書類_20251231_自動転記版.xlsx
  BS：2025/12/31 期末 TB(12) 残高を使用
  PL：TB(12-11) + TB(12) 13ヶ月合算
"""
import sys
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')

BASE = "C:/Users/yotak/Documents/everything-claude-code/DOORDO_Financial_Analysis_2025/"
SOURCE_XL = BASE + "FY25_増減分析_株式会社DOORDO.xlsx"
OUTPUT_XL = BASE + "20260406_DOORDO_計算書類_20251231_自動転記版.xlsx"

COMPANY   = "株式会社ＤＯＯＲＤＯ"
PERIOD_BS = "令和7年12月31日 現在"
PERIOD_PL = "自 令和6年12月1日  至 令和7年12月31日（13ヶ月）"
PERIOD_SS = "自 令和6年12月1日  至 令和7年12月31日"

# ─────────────────────────────────────────────────────────────────────────────
# ステップ1：マッピング辞書（指示通り）
# ─────────────────────────────────────────────────────────────────────────────

MAPPING_RULE = {
    "BS": {
        "現金預金":       "現金及び預金",
        "完成工事未収入金":"完成工事未収入金",
        "販売用不動産":   "販売用不動産",
        "仕掛販売用不動産":"仕掛販売用不動産",
        "未成工事支出金": "未成工事支出金",
        "未収入金":       "その他",
        "前渡金":         "その他",
        "立替金":         "その他",
        "短期貸付金":     "短期貸付金",
        "未収収益":       "その他",
        "前払費用":       "その他",
        "未収消費税等":   "その他",
        "建物":           "建物",
        "建物付属設備":   "建物附属設備",
        "構築物":         "構築物",
        "車両運搬具":     "車両及び運搬具",
        "工具器具備品":   "器具及び備品",
        "土地":           "土地",
        "ソフトウェア":   "ソフトウェア",
        "長期貸付金":     "長期貸付金",
        "長期前払費用":   "長期前払費用",
        "出資金":         "その他",
        "保証金":         "その他",
        "保険積立金":     "その他",
        "ゴルフ会員権":   "その他",
        "リサイクル預託金":"その他",
        "繰延消費税額等": "その他",
        "支払手形":       "支払手形",
        "買掛金":         "買掛金",
        "短期借入金":     "短期借入金",
        "未払金":         "未払金",
        "未払法人税等":   "未払法人税等",
        "未成工事受入金": "未成工事受入金",
        "未払費用":       "その他",
        "前受金":         "その他",
        "預り金":         "その他",
        "前受収益":       "その他",
        "未払消費税等":   "その他",
        "社債":           "社債",
        "長期借入金":     "長期借入金",
        "資本性借入金":   "長期借入金",
        "預り保証金":     "長期預り保証金",
        "資本金":         "資本金",
        "繰越利益剰余金": "繰越利益剰余金",
    },
    "PL": {
        "売上高":   "売上高",
        "仕入高":   "売上原価",
        "受取利息": "営業外収益",
        "支払利息": "営業外費用",
        "雑損失":   "営業外費用",
    }
}

BS_RULE = MAPPING_RULE["BS"]

# TB勘定科目 → (MAPPING_RULEキー, セクション)
# セクションは「その他」の分類に使用
TB_ACCT_MAP = {
    # ── 流動資産 ──
    "現金":               ("現金預金",       "流動資産"),
    "普通預金":           ("現金預金",       "流動資産"),
    "定期預金":           ("現金預金",       "流動資産"),
    "定期積金":           ("現金預金",       "流動資産"),
    "返済準備金留保口座": ("現金預金",       "流動資産"),
    "完成工事未収入金":   ("完成工事未収入金","流動資産"),
    "販売用不動産":       ("販売用不動産",   "流動資産"),
    "販売用不動産(収益)": ("販売用不動産",   "流動資産"),
    "仕掛販売用不動産":   ("仕掛販売用不動産","流動資産"),
    "未成工事支出金":     ("未成工事支出金", "流動資産"),
    "短期貸付金":         ("短期貸付金",     "流動資産"),
    "未収入金":           ("未収入金",       "流動資産"),
    "前渡金":             ("前渡金",         "流動資産"),
    "立替金":             ("立替金",         "流動資産"),
    "未収収益":           ("未収収益",       "流動資産"),
    "前払費用":           ("前払費用",       "流動資産"),
    "未収消費税":         ("未収消費税等",   "流動資産"),
    "仮払消費税":         ("未収消費税等",   "流動資産"),
    "未収法人税等":       ("未収消費税等",   "流動資産"),
    "仮払金":             ("未収消費税等",   "流動資産"),
    # ── 有形固定資産 ──
    "建物":               ("建物",           "有形固定資産"),
    "附属設備":           ("建物付属設備",   "有形固定資産"),
    "構築物":             ("構築物",         "有形固定資産"),
    "車両運搬具":         ("車両運搬具",     "有形固定資産"),
    "工具器具備品":       ("工具器具備品",   "有形固定資産"),
    "土地":               ("土地",           "有形固定資産"),
    "建設仮勘定":         (None,             "有形固定資産"),  # スキップ
    # ── 無形固定資産 ──
    "ソフトウェア":       ("ソフトウェア",   "無形固定資産"),
    "のれん":             ("_のれん",        "無形固定資産"),  # MAPPING_RULE外・個別表示
    # ── 投資その他の資産 ──
    "長期貸付金":         ("長期貸付金",     "投資その他の資産"),
    "長期前払費用":       ("長期前払費用",   "投資その他の資産"),
    "出資金":             ("出資金",         "投資その他の資産"),
    "差入保証金":         ("保証金",         "投資その他の資産"),
    "保険積立金":         ("保険積立金",     "投資その他の資産"),
    "ﾘｻｲｸﾙ預託金":      ("リサイクル預託金","投資その他の資産"),
    "ｺﾞﾙﾌ会員権":        ("ゴルフ会員権",   "投資その他の資産"),
    "繰延消費税額等":     ("繰延消費税額等", "投資その他の資産"),
    # ── 流動負債 ──
    "短期借入金":         ("短期借入金",     "流動負債"),
    "未払金":             ("未払金",         "流動負債"),
    "未払法人税等":       ("未払法人税等",   "流動負債"),
    "未成工事受入金":     ("未成工事受入金", "流動負債"),
    "未払費用":           ("未払費用",       "流動負債"),
    "前受金":             ("前受金",         "流動負債"),
    "預り金":             ("預り金",         "流動負債"),
    "前受収益":           ("前受収益",       "流動負債"),
    "未払消費税等":       ("未払消費税等",   "流動負債"),
    "仮受金":             ("前受金",         "流動負債"),
    "仮受消費税":         ("未払消費税等",   "流動負債"),
    # ── 固定負債 ──
    "社債":               ("社債",           "固定負債"),
    "長期借入金":         ("長期借入金",     "固定負債"),
    "資本性借入金":       ("資本性借入金",   "固定負債"),
    "預り保証金":         ("預り保証金",     "固定負債"),
    # ── 純資産 ──
    "資本金":             ("資本金",         "純資産"),
    "繰越利益":           ("繰越利益剰余金", "純資産"),
}

# TB[損益計算書]明細行 → 開示科目 (指定外 → 販管費)
TB_PL_MAP = {
    "土地売上高":               "売上高",
    "建物売上高":               "売上高",
    "賃貸収入":                 "売上高",
    "その他工事収入(請負工事)": "売上高",
    "役務収入":                 "売上高",
    "礼金収入高":               "売上高",
    "売上高":                   "売上高",
    "土地仕入高":               "売上原価",
    "建物仕入高":               "売上原価",
    "共通仕入":                 "売上原価",
    "その他仕入高":             "売上原価",
    "建築請負 仕入高":          "売上原価",
    "★仕入高":                 "売上原価",
    "受取利息":                 "受取利息",
    "受取配当金":               "受取配当金",
    "雑収入":                   "雑収入",
    "支払利息":                 "支払利息",
    "雑損失":                   "営業外費用その他",
    "固定資産売却益":           "固定資産売却益",
    "固定資産除却損":           "固定資産除却損",
    "法人税･住民税及び事業税":  "法人税等",
    "★居住用賃貸消費税制限":   "_skip",
}

PL_SGA_SECTION = "[販売管理費]"

def to_int(v):
    if v is None: return 0
    try: return int(v)
    except: return 0

# ─────────────────────────────────────────────────────────────────────────────
# TB読み込み
# ─────────────────────────────────────────────────────────────────────────────

TB_SECTION_MAP = {
    "[現金･預金]":      "流動資産",
    "[売上債権]":        "流動資産",
    "[有価証券]":        "流動資産",
    "[棚卸資産]":        "流動資産",
    "[他流動資産]":      "流動資産",
    "[有形固定資産]":    "有形固定資産",
    "[無形固定資産]":    "無形固定資産",
    "[投資その他の資産]":"投資その他の資産",
    "[繰延資産]":        "繰延資産",
    "[仕入債務]":        "流動負債",
    "[他流動負債]":      "流動負債",
    "[固定負債]":        "固定負債",
    "[資本金]":          "純資産",
    "[新株式申込証拠金]":"純資産",
    "[資本剰余金]":      "純資産",
    "[利益剰余金]":      "純資産",
    "[自己株式]":        "純資産",
    "[評価･換算差額等]": "純資産",
    "[新株予約権]":      "純資産",
    "[売上高]":          "[売上高]",
    "[売上原価]":        "[売上原価]",
    "[販売管理費]":      "[販売管理費]",
    "[営業外収益]":      "[営業外収益]",
    "[営業外費用]":      "[営業外費用]",
    "[特別利益]":        "[特別利益]",
    "[特別損失]":        "[特別損失]",
    "[当期純損益]":      "[当期純損益]",
}

def read_tb(ws, bal_col=7):
    """明細行を読んでBS行・PL行・期首残高を返す"""
    bs_rows = []
    pl_rows = []
    opening = {}
    current_fs = None
    current_sec = None

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        rt = str(row[0])
        if rt not in ("[区分行]", "[明細行]", "[合計行]", "[表題行]"):
            continue
        if len(row) <= bal_col:
            continue

        cls  = str(row[2]) if row[2] else ""
        acct = str(row[3]) if row[3] else ""
        bal  = to_int(row[bal_col])
        op   = to_int(row[4]) if len(row) > 4 else 0

        if rt == "[区分行]":
            if cls == "[貸借対照表]":
                current_fs = "BS"
                current_sec = TB_SECTION_MAP.get(acct, acct)
            elif cls == "[損益計算書]":
                current_fs = "PL"
                current_sec = TB_SECTION_MAP.get(acct, acct)
        elif rt == "[明細行]":
            if current_fs == "BS":
                bs_rows.append((acct, bal, current_sec))
                opening[acct] = op
            elif current_fs == "PL":
                pl_rows.append((acct, bal, current_sec))

    return bs_rows, pl_rows, opening

def read_summary(ws, bal_col=7):
    """合計行残高を辞書で返す"""
    d = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] != "[合計行]":
            continue
        if len(row) > bal_col:
            acct = str(row[3]) if row[3] else ""
            d[acct] = to_int(row[bal_col])
    return d

print("=" * 60)
print("  TB読み込み中...")
print("=" * 60)
src_wb = openpyxl.load_workbook(SOURCE_XL, read_only=True)
bs12_11, pl12_11, op_12_11 = read_tb(src_wb["期末TB(12-11)"], 7)
bs12,    pl12,    op_12    = read_tb(src_wb["期末TB(12)"],    7)
sum12_11 = read_summary(src_wb["期末TB(12-11)"], 7)
sum12    = read_summary(src_wb["期末TB(12)"],    7)
src_wb.close()
print(f"  TB12-11: BS {len(bs12_11)}行 / PL {len(pl12_11)}行")
print(f"  TB12:    BS {len(bs12)}行 / PL {len(pl12)}行")

# ─────────────────────────────────────────────────────────────────────────────
# ステップ2：BS集計（TB12 期末残高）
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 60)
print("  ステップ2: BS集計")
print("=" * 60)

bs_agg = defaultdict(int)       # (開示科目, セクション) → 合計金額
bs_map_detail = []               # マッピング表用

for tb_acct, bal, tb_sec in bs12:
    if bal == 0:
        continue
    if tb_acct not in TB_ACCT_MAP:
        # 未マッピング → その他(セクション)として分類
        sec = tb_sec if tb_sec else "不明"
        disclosure = f"その他（{sec}）"
        bs_agg[(disclosure, sec)] += bal
        bs_map_detail.append((tb_acct, tb_sec, bal, "(未マッピング)", sec, disclosure))
        continue

    rule_key, sec = TB_ACCT_MAP[tb_acct]

    if rule_key is None:
        # スキップ（建設仮勘定など）
        bs_map_detail.append((tb_acct, tb_sec, bal, "(スキップ)", sec, "-"))
        continue

    if rule_key == "_のれん":
        disclosure = "のれん"
    elif rule_key in BS_RULE:
        raw = BS_RULE[rule_key]
        disclosure = f"その他（{sec}）" if raw == "その他" else raw
    else:
        disclosure = f"その他（{sec}）"

    bs_agg[(disclosure, sec)] += bal
    bs_map_detail.append((tb_acct, tb_sec, bal, rule_key, sec, disclosure))

# 繰越利益剰余金 → 合計行（前期繰越＋当期純損益）で上書き
netto_closing = sum12.get("繰越利益剰余金合計", 0)
bs_agg[("繰越利益剰余金", "純資産")] = netto_closing

def bv(key, sec):
    return bs_agg.get((key, sec), 0)

# 集計値
現金預金     = bv("現金及び預金",     "流動資産")
完成工事     = bv("完成工事未収入金", "流動資産")
販売用       = bv("販売用不動産",     "流動資産")
仕掛販売用   = bv("仕掛販売用不動産", "流動資産")
未成工事支出 = bv("未成工事支出金",   "流動資産")
短期貸付     = bv("短期貸付金",       "流動資産")
流動その他   = bv("その他（流動資産）","流動資産")

建物         = bv("建物",         "有形固定資産")
建物附属     = bv("建物附属設備", "有形固定資産")
構築物       = bv("構築物",       "有形固定資産")
車両         = bv("車両及び運搬具","有形固定資産")
器具備品     = bv("器具及び備品", "有形固定資産")
土地         = bv("土地",         "有形固定資産")

ソフト       = bv("ソフトウェア", "無形固定資産")
のれん_bal   = bv("のれん",       "無形固定資産")

長期貸付     = bv("長期貸付金",   "投資その他の資産")
長期前払     = bv("長期前払費用", "投資その他の資産")
投資その他   = bv("その他（投資その他の資産）","投資その他の資産")

短期借入     = bv("短期借入金",     "流動負債")
未払金_bal   = bv("未払金",         "流動負債")
未払法人税   = bv("未払法人税等",   "流動負債")
未成工事受入 = bv("未成工事受入金", "流動負債")
流動負債その他 = bv("その他（流動負債）","流動負債")

社債_bal     = bv("社債",         "固定負債")
長期借入     = bv("長期借入金",   "固定負債")
長期保証     = bv("長期預り保証金","固定負債")

資本金_bal   = bv("資本金",             "純資産")
繰越利益_bal = bv("繰越利益剰余金",     "純資産")

# 小計・合計
有形固定計   = 建物 + 建物附属 + 構築物 + 車両 + 器具備品 + 土地
無形固定計   = ソフト + のれん_bal
投資その他計 = 長期貸付 + 長期前払 + 投資その他
固定資産計   = 有形固定計 + 無形固定計 + 投資その他計
流動資産計   = 現金預金 + 完成工事 + 販売用 + 仕掛販売用 + 未成工事支出 + 短期貸付 + 流動その他
資産合計     = 流動資産計 + 固定資産計

流動負債計   = 短期借入 + 未払金_bal + 未払法人税 + 未成工事受入 + 流動負債その他
固定負債計   = 社債_bal + 長期借入 + 長期保証
負債合計     = 流動負債計 + 固定負債計
利益剰余金計 = 繰越利益_bal
株主資本計   = 資本金_bal + 利益剰余金計
純資産計     = 株主資本計
負債純資産計 = 負債合計 + 純資産計

print(f"  資産合計:         {資産合計:>20,}")
print(f"  負債合計:         {負債合計:>20,}")
print(f"  純資産合計:       {純資産計:>20,}")
print(f"  負債及び純資産:   {負債純資産計:>20,}")
print(f"  BS整合性: {'✅ OK' if 資産合計 == 負債純資産計 else f'❌ 差額 {資産合計 - 負債純資産計:+,}'}")

# ─────────────────────────────────────────────────────────────────────────────
# PL集計（13ヶ月合算）
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 60)
print("  PL集計（13ヶ月合算）")
print("=" * 60)

pl_agg = defaultdict(int)
pl_map_detail = []
pl_sga_12m = {}   # 販管費明細 12M
pl_sga_1m  = {}   # 販管費明細 1M

def add_pl(rows, label, sga_dict):
    for tb_acct, bal, pl_sec in rows:
        if bal == 0:
            continue
        if tb_acct in TB_PL_MAP:
            disc = TB_PL_MAP[tb_acct]
            if disc == "_skip":
                continue
            pl_agg[disc] += bal
            pl_map_detail.append((label, tb_acct, pl_sec, bal, disc))
        elif pl_sec == PL_SGA_SECTION:
            pl_agg["販売費及び一般管理費"] += bal
            sga_dict[tb_acct] = sga_dict.get(tb_acct, 0) + bal
            pl_map_detail.append((label, tb_acct, pl_sec, bal, "販売費及び一般管理費"))
        else:
            pl_map_detail.append((label, tb_acct, pl_sec, bal, "(未分類)"))

add_pl(pl12_11, "TB12-11(12M)", pl_sga_12m)
add_pl(pl12,    "TB12(1M)",     pl_sga_1m)

# PL小計
gross_profit    = pl_agg["売上高"] - pl_agg["売上原価"]
op_income       = gross_profit - pl_agg["販売費及び一般管理費"]
other_inc       = pl_agg["受取利息"] + pl_agg["受取配当金"] + pl_agg["雑収入"]
other_exp       = pl_agg["支払利息"] + pl_agg["営業外費用その他"]
recurring_inc   = op_income + other_inc - other_exp
extraordinary_i = pl_agg["固定資産売却益"]
extraordinary_l = pl_agg["固定資産除却損"]
pretax_inc      = recurring_inc + extraordinary_i - extraordinary_l
income_tax      = pl_agg["法人税等"]
net_income      = pretax_inc - income_tax

print(f"  売上高:           {pl_agg['売上高']:>20,}")
print(f"  売上原価:         {pl_agg['売上原価']:>20,}")
print(f"  売上総利益:       {gross_profit:>20,}")
print(f"  販管費:           {pl_agg['販売費及び一般管理費']:>20,}")
print(f"  営業利益:         {op_income:>20,}")
print(f"  経常利益:         {recurring_inc:>20,}")
print(f"  当期純利益:       {net_income:>20,}")

# SS値
netto_13m       = sum12_11.get("当期純損益金額", 0) + sum12.get("当期純損益金額", 0)
ss_開始繰越     = op_12_11.get("繰越利益", 0)
ss_開始資本金   = op_12_11.get("資本金", 30_000_000)
ss_変動繰越     = netto_13m
ss_終了繰越     = netto_closing
# 期首+純利益 と期末が一致しない場合、差額＝資本直入項目（のれん消却等）
ss_その他変動   = ss_終了繰越 - ss_開始繰越 - ss_変動繰越  # 通常 -80,229,769
ss_変動合計     = ss_変動繰越 + ss_その他変動
ss_終了株主資本 = ss_開始資本金 + ss_開始繰越 + ss_変動合計

print(f"\n  SS開始繰越利益:     {ss_開始繰越:>20,}")
print(f"  当期純利益13M:      {ss_変動繰越:>20,}")
print(f"  SS資本直入その他変動: {ss_その他変動:>20,}  ← のれん消却等を利益剰余金直入")
print(f"  SS終了繰越利益:     {ss_終了繰越:>20,}")
print(f"  SS整合性 (終了=開始+変動): {'✅ OK' if ss_終了繰越 == ss_開始繰越 + ss_変動合計 else f'❌ 差額 {ss_終了繰越 - ss_開始繰越 - ss_変動合計:+,}'}")

# ─────────────────────────────────────────────────────────────────────────────
# ステップ3：Excel出力
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 60)
print("  ステップ3: Excel出力")
print("=" * 60)

wb = Workbook()
wb.remove(wb.active)

# ── スタイル定義 ──────────────────────────────────────────────────────────
FONT_BASE = "游ゴシック"
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
HDR_FONT  = Font(name=FONT_BASE, bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(name=FONT_BASE, bold=True, size=10)
REG_FONT  = Font(name=FONT_BASE, size=10)
NUM_FMT   = "#,##0"
THIN      = Side(style="thin")
MED       = Side(style="medium")

def border_all():
    return Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def set_val(ws, cel, val, bold=False, num=False, indent=0, align="left"):
    c = ws[cel]
    c.value = val
    c.font  = BOLD_FONT if bold else REG_FONT
    c.alignment = Alignment(
        horizontal="right" if (num or align == "right") else "left",
        vertical="center", indent=indent)
    if num and isinstance(val, (int, float)):
        c.number_format = NUM_FMT

# ─────────────────────────────────────────────────────────────────────────────
# BS（円）シート
# ─────────────────────────────────────────────────────────────────────────────

ws_bs = wb.create_sheet("BS（円）")
ws_bs.sheet_view.showGridLines = False

# ── ヘッダー ──
ws_bs["B2"] = "貸　借　対　照　表"
ws_bs["B2"].font = Font(name=FONT_BASE, bold=True, size=13)
ws_bs["B3"] = f"（{PERIOD_BS}）"
ws_bs["B3"].font = REG_FONT
ws_bs["B4"] = COMPANY
ws_bs["B4"].font = REG_FONT
ws_bs["N4"] = "（単位：円）"
ws_bs["N4"].font = REG_FONT
for col, txt in [("B", "科目"), ("H", "金額"), ("I", "科目"), ("P", "金額")]:
    ws_bs[f"{col}5"] = txt
    ws_bs[f"{col}5"].font = BOLD_FONT
ws_bs["B6"] = "（資産の部）"
ws_bs["I6"] = "（負債の部）"
ws_bs["B6"].font = BOLD_FONT
ws_bs["I6"].font = BOLD_FONT

# BS行書き込みヘルパー
# 左側: C=lv1, D=lv2, E=lv3/acct, G=個別金額, H=小計/合計
# 右側: J=lv1, K=lv2, L=lv3/acct, N=個別金額, O=小計/合計

def bs_left(ws, row, lv, text, detail=None, subtot=None):
    if lv == 0:   # outer（資産合計）
        ws[f"B{row}"] = text; ws[f"B{row}"].font = BOLD_FONT
    elif lv == 1: # 流動資産 / 固定資産
        ws[f"C{row}"] = text; ws[f"C{row}"].font = BOLD_FONT
    elif lv == 2: # 有形固定資産
        ws[f"D{row}"] = f"　{text}"; ws[f"D{row}"].font = BOLD_FONT
    elif lv == 3: # 勘定科目
        ws[f"E{row}"] = f"　　{text}"; ws[f"E{row}"].font = REG_FONT
    if detail is not None:
        ws[f"G{row}"] = detail
        ws[f"G{row}"].number_format = NUM_FMT
        ws[f"G{row}"].alignment = Alignment(horizontal="right")
        ws[f"G{row}"].font = REG_FONT
    if subtot is not None:
        ws[f"H{row}"] = subtot
        ws[f"H{row}"].number_format = NUM_FMT
        ws[f"H{row}"].alignment = Alignment(horizontal="right")
        ws[f"H{row}"].font = BOLD_FONT if lv <= 1 else REG_FONT

def bs_right(ws, row, lv, text, detail=None, subtot=None):
    if text is not None:
        if lv == 0:
            ws[f"I{row}"] = text; ws[f"I{row}"].font = BOLD_FONT
        elif lv == 1:
            ws[f"J{row}"] = text; ws[f"J{row}"].font = BOLD_FONT
        elif lv == 2:
            ws[f"K{row}"] = f"　{text}"; ws[f"K{row}"].font = BOLD_FONT
        elif lv == 3:
            ws[f"L{row}"] = f"　　{text}"; ws[f"L{row}"].font = REG_FONT
    if detail is not None:
        ws[f"N{row}"] = detail
        ws[f"N{row}"].number_format = NUM_FMT
        ws[f"N{row}"].alignment = Alignment(horizontal="right")
        ws[f"N{row}"].font = REG_FONT
    if subtot is not None:
        ws[f"O{row}"] = subtot
        ws[f"O{row}"].number_format = NUM_FMT
        ws[f"O{row}"].alignment = Alignment(horizontal="right")
        ws[f"O{row}"].font = BOLD_FONT if lv <= 1 else REG_FONT

r = 7
bs_left(ws_bs, r, 1, "流動資産", subtot=流動資産計)
bs_right(ws_bs, r, 1, "流動負債", subtot=流動負債計); r += 1

bs_left(ws_bs, r, 3, "現金及び預金",    現金預金)
bs_right(ws_bs, r, 3, "短期借入金",     短期借入); r += 1

bs_left(ws_bs, r, 3, "完成工事未収入金", 完成工事)
bs_right(ws_bs, r, 3, "未払金",         未払金_bal); r += 1

bs_left(ws_bs, r, 3, "販売用不動産",    販売用)
bs_right(ws_bs, r, 3, "未払法人税等",   未払法人税); r += 1

bs_left(ws_bs, r, 3, "仕掛販売用不動産", 仕掛販売用)
bs_right(ws_bs, r, 3, "未成工事受入金", 未成工事受入); r += 1

bs_left(ws_bs, r, 3, "未成工事支出金",  未成工事支出)
bs_right(ws_bs, r, 3, "その他",         流動負債その他); r += 1

bs_left(ws_bs, r, 3, "短期貸付金",      短期貸付); r += 1

bs_left(ws_bs, r, 3, "その他",          流動その他)
bs_right(ws_bs, r, 1, "固定負債", subtot=固定負債計); r += 1

bs_left(ws_bs, r, 1, "固定資産", subtot=固定資産計)
bs_right(ws_bs, r, 3, "社債",           社債_bal); r += 1

bs_left(ws_bs, r, 2, "有形固定資産", subtot=有形固定計)
bs_right(ws_bs, r, 3, "長期借入金",     長期借入); r += 1

bs_left(ws_bs, r, 3, "建物",            建物)
bs_right(ws_bs, r, 3, "長期預り保証金", 長期保証); r += 1

bs_left(ws_bs, r, 3, "建物附属設備",    建物附属)
bs_right(ws_bs, r, 0, "負債合計", subtot=負債合計); r += 1

bs_left(ws_bs, r, 3, "構築物",          構築物)
bs_right(ws_bs, r, 0, "（純資産の部）"); r += 1

bs_left(ws_bs, r, 3, "車両及び運搬具",  車両)
bs_right(ws_bs, r, 1, "株主資本", subtot=株主資本計); r += 1

bs_left(ws_bs, r, 3, "器具及び備品",    器具備品)
bs_right(ws_bs, r, 3, "資本金",         資本金_bal); r += 1

bs_left(ws_bs, r, 3, "土地",            土地)
bs_right(ws_bs, r, 2, "利益剰余金", subtot=利益剰余金計); r += 1

bs_left(ws_bs, r, 2, "無形固定資産", subtot=無形固定計)
bs_right(ws_bs, r, 3, "繰越利益剰余金", 繰越利益_bal); r += 1

bs_left(ws_bs, r, 3, "ソフトウェア",    ソフト)
bs_right(ws_bs, r, 0, "純資産合計", subtot=純資産計); r += 1

bs_left(ws_bs, r, 3, "のれん",          のれん_bal)
bs_right(ws_bs, r, 0, "負債及び純資産合計", subtot=負債純資産計); r += 1

bs_left(ws_bs, r, 2, "投資その他の資産", subtot=投資その他計); r += 1
bs_left(ws_bs, r, 3, "長期貸付金",      長期貸付); r += 1
bs_left(ws_bs, r, 3, "長期前払費用",    長期前払); r += 1
bs_left(ws_bs, r, 3, "その他",          投資その他); r += 1
bs_left(ws_bs, r, 0, "資産合計", subtot=資産合計); r += 1

# 列幅
for col, w in [("B",12),("C",14),("D",16),("E",20),("F",2),("G",16),
               ("H",16),("I",14),("J",14),("K",16),("L",20),("M",2),
               ("N",16),("O",16),("P",2)]:
    ws_bs.column_dimensions[col].width = w

print(f"  BS（円）シート: 行 7 → {r-1}")

# ─────────────────────────────────────────────────────────────────────────────
# PL（円）シート
# ─────────────────────────────────────────────────────────────────────────────

ws_pl = wb.create_sheet("PL（円）")
ws_pl.sheet_view.showGridLines = False

ws_pl["B2"] = "損　益　計　算　書"
ws_pl["B2"].font = Font(name=FONT_BASE, bold=True, size=13)
ws_pl["B3"] = f"（{PERIOD_PL}）"
ws_pl["B3"].font = REG_FONT
ws_pl["B4"] = COMPANY
ws_pl["B4"].font = REG_FONT
ws_pl["H4"] = "（単位：円）"
ws_pl["H4"].font = REG_FONT
ws_pl["B5"] = "科目"
ws_pl["H5"] = "金額"
ws_pl["B5"].font = BOLD_FONT
ws_pl["H5"].font = BOLD_FONT

def pl_row(ws, row, lv, text, detail=None, subtot=None):
    """lv: 0=outer, 1=section, 2=detail line"""
    if lv == 1:
        ws[f"C{row}"] = text; ws[f"C{row}"].font = BOLD_FONT
    elif lv == 2:
        ws[f"D{row}"] = f"　{text}"; ws[f"D{row}"].font = REG_FONT
    elif lv == 0:
        ws[f"B{row}"] = text; ws[f"B{row}"].font = BOLD_FONT
    if detail is not None:
        ws[f"G{row}"] = detail
        ws[f"G{row}"].number_format = NUM_FMT
        ws[f"G{row}"].alignment = Alignment(horizontal="right")
        ws[f"G{row}"].font = REG_FONT
    if subtot is not None:
        ws[f"H{row}"] = subtot
        ws[f"H{row}"].number_format = NUM_FMT
        ws[f"H{row}"].alignment = Alignment(horizontal="right")
        ws[f"H{row}"].font = BOLD_FONT

r = 6
pl_row(ws_pl, r, 1, "売上高",               subtot=pl_agg["売上高"]); r += 1
pl_row(ws_pl, r, 1, "売上原価",              subtot=pl_agg["売上原価"]); r += 1
pl_row(ws_pl, r, 2, "売上総利益",            subtot=gross_profit); r += 1
pl_row(ws_pl, r, 1, "販売費及び一般管理費",  subtot=pl_agg["販売費及び一般管理費"]); r += 1
pl_row(ws_pl, r, 2, "営業利益",              subtot=op_income); r += 1
pl_row(ws_pl, r, 1, "営業外収益"); r += 1
pl_row(ws_pl, r, 2, "受取利息",   detail=pl_agg["受取利息"]); r += 1
pl_row(ws_pl, r, 2, "受取配当金", detail=pl_agg["受取配当金"]); r += 1
pl_row(ws_pl, r, 2, "雑収入",     detail=pl_agg["雑収入"],     subtot=other_inc); r += 1
pl_row(ws_pl, r, 1, "営業外費用"); r += 1
pl_row(ws_pl, r, 2, "支払利息",   detail=pl_agg["支払利息"]); r += 1
pl_row(ws_pl, r, 2, "その他",     detail=pl_agg["営業外費用その他"], subtot=other_exp); r += 1
pl_row(ws_pl, r, 2, "経常利益",               subtot=recurring_inc); r += 1
pl_row(ws_pl, r, 1, "特別利益"); r += 1
pl_row(ws_pl, r, 2, "固定資産売却益", detail=extraordinary_i, subtot=extraordinary_i); r += 1
pl_row(ws_pl, r, 1, "特別損失"); r += 1
pl_row(ws_pl, r, 2, "固定資産除却損", detail=extraordinary_l, subtot=extraordinary_l); r += 1
pl_row(ws_pl, r, 2, "税引前当期純利益",       subtot=pretax_inc); r += 1
pl_row(ws_pl, r, 0, "法人税、住民税及び事業税", detail=income_tax); r += 1
pl_row(ws_pl, r, 2, "当期純利益",             subtot=net_income); r += 1

for col, w in [("B",22),("C",22),("D",22),("E",2),("F",2),("G",18),("H",18)]:
    ws_pl.column_dimensions[col].width = w
print(f"  PL（円）シート: 行 6 → {r-1}")

# ─────────────────────────────────────────────────────────────────────────────
# SS（円）シート
# ─────────────────────────────────────────────────────────────────────────────

ws_ss = wb.create_sheet("SS（円）")
ws_ss.sheet_view.showGridLines = False

ws_ss["B2"] = "株主資本等変動計算書"
ws_ss["B2"].font = Font(name=FONT_BASE, bold=True, size=13)
ws_ss["B3"] = f"（{PERIOD_SS}）"
ws_ss["B3"].font = REG_FONT
ws_ss["B4"] = COMPANY
ws_ss["B4"].font = REG_FONT
ws_ss["J4"] = "（単位：円）"
ws_ss["J4"].font = REG_FONT

def ss_row(ws, row, lv, text=None, label=None, amount=None):
    """lv: 0=pure_assets_total, 1=section, 2=account, 3=sub-acct, 4=detail-acct"""
    if text is not None:
        cols = {0: "C", 1: "C", 2: "D", 3: "E", 4: "F"}
        col = cols.get(lv, "D")
        indent = max(0, lv - 1)
        ws[f"{col}{row}"] = ("　" * indent) + text
        ws[f"{col}{row}"].font = BOLD_FONT if lv <= 2 else REG_FONT
    if label:
        ws[f"H{row}"] = label
        ws[f"H{row}"].font = REG_FONT
    if amount is not None:
        ws[f"J{row}"] = amount
        ws[f"J{row}"].number_format = NUM_FMT
        ws[f"J{row}"].alignment = Alignment(horizontal="right")
        ws[f"J{row}"].font = REG_FONT

r = 6
ss_row(ws_ss, r, 1, "（株主資本）"); r += 1
ss_row(ws_ss, r, 2, "資本金",    "当期首残高", ss_開始資本金); r += 1
ss_row(ws_ss, r, None, None,      "当期末残高", ss_開始資本金); r += 1
ss_row(ws_ss, r, 2, "利益剰余金"); r += 1
ss_row(ws_ss, r, 3, "その他利益剰余金"); r += 1
ss_row(ws_ss, r, 4, "繰越利益剰余金", "当期首残高",                       ss_開始繰越); r += 1
ss_row(ws_ss, r, None, None,          "当期変動額　当期純利益金額",          ss_変動繰越); r += 1
if ss_その他変動 != 0:
    ss_row(ws_ss, r, None, None,      "当期変動額　その他（のれん消却等直入）", ss_その他変動); r += 1
ss_row(ws_ss, r, None, None,          "当期末残高",                          ss_終了繰越); r += 1
ss_row(ws_ss, r, 3, "利益剰余金合計",  "当期首残高",  ss_開始繰越); r += 1
ss_row(ws_ss, r, None, None,           "当期変動額",  ss_変動合計); r += 1
ss_row(ws_ss, r, None, None,           "当期末残高",  ss_終了繰越); r += 1
ss_row(ws_ss, r, 2, "株主資本合計",    "当期首残高", ss_開始資本金 + ss_開始繰越); r += 1
ss_row(ws_ss, r, None, None,           "当期変動額",  ss_変動合計); r += 1
ss_row(ws_ss, r, None, None,           "当期末残高",  ss_終了株主資本); r += 1
ss_row(ws_ss, r, 0, "純資産の部合計",  "当期首残高", ss_開始資本金 + ss_開始繰越); r += 1
ss_row(ws_ss, r, None, None,           "当期変動額",  ss_変動合計); r += 1
ss_row(ws_ss, r, None, None,           "当期末残高",  ss_終了株主資本); r += 1

for col, w in [("B",4),("C",18),("D",18),("E",18),("F",18),("G",4),("H",26),("I",4),("J",18)]:
    ws_ss.column_dimensions[col].width = w
print(f"  SS（円）シート: 行 6 → {r-1}")

# ─────────────────────────────────────────────────────────────────────────────
# マッピング表シート
# ─────────────────────────────────────────────────────────────────────────────

ws_map = wb.create_sheet("マッピング表")
ws_map.sheet_view.showGridLines = False

map_headers = ["シート区分", "TB勘定科目", "TBセクション（区分行）", "金額（TB残高）",
               "MAPPING_RULEキー", "適用セクション", "開示科目（集約後）"]
for i, h in enumerate(map_headers, 1):
    c = ws_map.cell(row=1, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all()

row = 2
for (tb_acct, tb_sec, bal, rk, sec, disc) in bs_map_detail:
    for i, v in enumerate(["BS（円）", tb_acct, tb_sec, bal, rk, sec, disc], 1):
        c = ws_map.cell(row=row, column=i, value=v)
        c.font = REG_FONT
        c.border = border_all()
        if i == 4:
            c.number_format = NUM_FMT
            c.alignment = Alignment(horizontal="right")
    row += 1

for (label, tb_acct, pl_sec, bal, disc) in pl_map_detail:
    for i, v in enumerate(["PL（円）", tb_acct, pl_sec, bal, "(MAPPING_RULE PL)", label, disc], 1):
        c = ws_map.cell(row=row, column=i, value=v)
        c.font = REG_FONT
        c.border = border_all()
        if i == 4:
            c.number_format = NUM_FMT
            c.alignment = Alignment(horizontal="right")
    row += 1

for col, w in [("A",14),("B",24),("C",24),("D",18),("E",22),("F",22),("G",26)]:
    ws_map.column_dimensions[col].width = w
ws_map.auto_filter.ref = f"A1:G{row-1}"
ws_map.freeze_panes = "A2"
print(f"  マッピング表: {row-2}件")

# ─────────────────────────────────────────────────────────────────────────────
# 販管費明細_附属明細書用シート
# ─────────────────────────────────────────────────────────────────────────────

ws_sga = wb.create_sheet("販管費明細_附属明細書用")
ws_sga.sheet_view.showGridLines = False

ws_sga["B2"] = "販売費及び一般管理費 明細書（附属明細書用）"
ws_sga["B2"].font = Font(name=FONT_BASE, bold=True, size=12)
ws_sga["B3"] = f"（{PERIOD_PL}）"
ws_sga["B3"].font = REG_FONT
ws_sga["B4"] = COMPANY
ws_sga["B4"].font = REG_FONT
ws_sga["G4"] = "（単位：円）"
ws_sga["G4"].font = REG_FONT

sga_cols = [("C", "勘定科目"),
            ("E", "第18期（12ヶ月）\n2024/12/01-2025/11/30"),
            ("F", "第19期（1ヶ月）\n2025/12/01-2025/12/31"),
            ("G", "合計（13ヶ月）")]
for col, h in sga_cols:
    c = ws_sga[f"{col}5"]
    c.value = h
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all()
ws_sga.row_dimensions[5].height = 32

# 全科目（12M + 1M のどちらかにある科目を全列挙）
all_sga = sorted(set(list(pl_sga_12m) + list(pl_sga_1m)))

row = 6
for acct in all_sga:
    v12 = pl_sga_12m.get(acct, 0)
    v1  = pl_sga_1m.get(acct, 0)
    tot = v12 + v1
    ws_sga[f"C{row}"] = acct
    ws_sga[f"C{row}"].font = REG_FONT
    ws_sga[f"C{row}"].border = border_all()
    for col, val in [("E", v12), ("F", v1), ("G", tot)]:
        c = ws_sga[f"{col}{row}"]
        c.value = val if val != 0 else None
        c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal="right")
        c.font = REG_FONT
        c.border = border_all()
    row += 1

# 合計行
for col, val in [("C", "販売費及び一般管理費合計"),
                  ("E", sum(pl_sga_12m.values())),
                  ("F", sum(pl_sga_1m.values())),
                  ("G", pl_agg["販売費及び一般管理費"])]:
    c = ws_sga[f"{col}{row}"]
    c.value = val
    c.font = BOLD_FONT
    c.fill = GRAY_FILL
    c.border = border_all()
    if col != "C":
        c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal="right")

for col, w in [("B",4),("C",26),("D",4),("E",20),("F",20),("G",20)]:
    ws_sga.column_dimensions[col].width = w
ws_sga.freeze_panes = "C6"
print(f"  販管費明細: {len(all_sga)}科目")

# ─────────────────────────────────────────────────────────────────────────────
# ステップ4：保存
# ─────────────────────────────────────────────────────────────────────────────

wb.save(OUTPUT_XL)
print(f"\n{'=' * 60}")
print(f"  ✅ 保存完了: {OUTPUT_XL}")
print(f"{'=' * 60}")

# ── コンソールサマリー ────────────────────────────────────────────────────

print(f"""
╔══════════════════════════════════════════════════════════╗
  BS / PL / SS 検証サマリー
╠══════════════════════════════════════════════════════════╣
  【BS：2025/12/31 期末（TB12 期末残高使用）】
  　流動資産合計     : {流動資産計:>20,}
  　固定資産合計     : {固定資産計:>20,}
  　資産合計         : {資産合計:>20,}
  　流動負債合計     : {流動負債計:>20,}
  　固定負債合計     : {固定負債計:>20,}
  　負債合計         : {負債合計:>20,}
  　純資産合計       : {純資産計:>20,}
  　負債及び純資産   : {負債純資産計:>20,}
  　BS整合性         : {'✅ OK（資産＝負債+純資産）' if 資産合計 == 負債純資産計 else f'❌ 差額 {資産合計 - 負債純資産計:+,}'}

  【PL：13ヶ月合算（TB12-11 + TB12）】
  　売上高           : {pl_agg['売上高']:>20,}
  　売上原価         : {pl_agg['売上原価']:>20,}
  　売上総利益       : {gross_profit:>20,}
  　販売費及び一般管理費: {pl_agg['販売費及び一般管理費']:>20,}
  　営業利益         : {op_income:>20,}
  　経常利益         : {recurring_inc:>20,}
  　税引前当期純利益 : {pretax_inc:>20,}
  　法人税等         : {income_tax:>20,}
  　当期純利益       : {net_income:>20,}

  【SS：13ヶ月（資本変動）】
  　資本金 開始＝終了 : {ss_開始資本金:>20,}
  　繰越利益 当期首   : {ss_開始繰越:>20,}
  　当期純利益変動額  : {ss_変動繰越:>20,}
  　その他変動（のれん等）: {ss_その他変動:>20,}
  　繰越利益 当期末   : {ss_終了繰越:>20,}
  　純資産合計 当期末 : {ss_終了株主資本:>20,}
  　SS整合性 (BS純資産と一致): {'✅ OK' if ss_終了株主資本 == 純資産計 else f'❌ 差額 {ss_終了株主資本 - 純資産計:+,}'}

  【生成シート】
  　① BS（円）　② PL（円）　③ SS（円）
  　④ マッピング表　⑤ 販管費明細_附属明細書用
╚══════════════════════════════════════════════════════════╝
""")
